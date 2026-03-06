"""
===========================================================
  DESCARGA RELACIONAL - FISCALÍA DE CHILE
  Delitos Ingresados 2014-2025 (estructura relacional)
===========================================================

Genera un CSV con: anno, mes, delito, comuna, fiscalia_regional, cantidad

Estrategia:
  1. Descarga listado de comunas y fiscalías (para parseo)
  2. Descarga cruce comuna-fiscalía → mapeo (fiscalía principal por comuna)
  3. Descarga cruce delito-comuna por año/mes
  4. Consolida todo en un CSV relacional

Uso:
  py descargar_relacional.py          # Ejecución completa (2014-2025)
  py descargar_relacional.py --test   # Modo test (2 meses de 2025)
"""

import requests
import os
import sys
import csv
import time
import json
import logging
import argparse
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl

# ======================================================
# CONFIGURACIÓN
# ======================================================

BASE_URL = "https://app-backend-reportabilidad-est-dev-eastus-001.azurewebsites.net/Descarga/GenerarDocumento"

FIXED_PARAMS = {
    "COD_REGION": "ALL", "COD_FISCALIA": "ALL",
    "GLS_TIPO_IMPUTADO": "ALL", "COD_MARCA_VIF": "ALL",
    "MARCA_RPA": "ALL", "COD_MARCA_ACD": "ALL",
    "COD_MARCA_ACD_FLAGRANCIA": "ALL", "COD_FAMILIADEL": "ALL",
    "COD_MATERIA": "ALL", "GLS_FAMILIADELVIF": "ALL",
    "TIPO_TERMINO": "ALL", "TIPO_GRUPO": "ALL",
    "MENOR_DE_EDAD": "ALL", "TIP_SEXO": "ALL",
    "TIP_PERSONA": "ALL", "COD_PAIS": "ALL",
    "COD_PARENTESCO": "ALL", "COD_REGION_OCURRENCIA": "ALL",
    "COD_COMUNA": "ALL", "GLS_PROCEDIMIENTO": "ALL",
    "Tabular_Columna": "TOTAL",
    "Tematica": "'Tabla de medidas'[Delitos_Ingresados]"
}

ANNO_INICIO = 2014
ANNO_FIN = 2025
MESES = list(range(1, 13))

PAUSA_ENTRE_DESCARGAS = 2
MAX_REINTENTOS = 3
TIMEOUT_DESCARGA = 180

# ======================================================
# FUNCIONES AUXILIARES
# ======================================================

def setup_logging(log_dir):
    os.makedirs(log_dir, exist_ok=True)
    log_file = os.path.join(log_dir, f"relacional_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(log_file, encoding="utf-8")
        ]
    )
    return log_file


def cargar_progreso(progress_file):
    if os.path.exists(progress_file):
        with open(progress_file, "r", encoding="utf-8") as f:
            return set(json.load(f))
    return set()


def guardar_progreso(progress_file, completados):
    with open(progress_file, "w", encoding="utf-8") as f:
        json.dump(list(completados), f)


def descargar_excel(params, filepath):
    """Descarga un archivo Excel con reintentos y backoff."""
    for intento in range(1, MAX_REINTENTOS + 1):
        try:
            response = requests.get(BASE_URL, params=params, timeout=TIMEOUT_DESCARGA)
            if response.status_code == 200 and len(response.content) > 100:
                with open(filepath, "wb") as f:
                    f.write(response.content)
                return True
            else:
                logging.warning(f"  HTTP {response.status_code} ({len(response.content)} bytes), "
                              f"intento {intento}/{MAX_REINTENTOS}")
        except requests.exceptions.Timeout:
            logging.warning(f"  Timeout, intento {intento}/{MAX_REINTENTOS}")
        except requests.exceptions.ConnectionError:
            logging.warning(f"  Error conexión, intento {intento}/{MAX_REINTENTOS}")
        except Exception as e:
            logging.warning(f"  Error: {e}, intento {intento}/{MAX_REINTENTOS}")
        if intento < MAX_REINTENTOS:
            time.sleep(5 * intento)
    return False


def leer_excel_columna1(filepath):
    """Lee un Excel y devuelve lista de (valor_col1, valor_col2) sin header."""
    datos = []
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if row[0] is None and not header_found:
            continue
        if not header_found:
            header_found = True
            continue
        if row[0] is not None:
            val1 = str(row[0]).strip()
            try:
                val2 = int(row[1]) if row[1] is not None else 0
            except (ValueError, TypeError):
                val2 = str(row[1]).strip() if row[1] else ""
            datos.append((val1, val2))
    wb.close()
    return datos


# ======================================================
# PASO 1: Obtener lista de comunas (para parseo de cruces)
# ======================================================

def obtener_comunas(base_dir):
    """Descarga o carga del caché la lista de todas las comunas."""
    cache = os.path.join(base_dir, "cache_comunas.json")
    if os.path.exists(cache):
        with open(cache, "r", encoding="utf-8") as f:
            comunas = json.load(f)
        logging.info(f"Comunas cargadas de caché: {len(comunas)}")
        return comunas

    logging.info("Descargando listado de comunas...")
    params = FIXED_PARAMS.copy()
    params.update({"ANNO": "2025", "MES_NOMBRE": "ALL",
                   "Tabular_Fila": "'GOLD DIM_COMUNA'[GLS_COMUNA]"})
    fp = os.path.join(base_dir, "_tmp_comunas.xlsx")
    if not descargar_excel(params, fp):
        return []
    
    datos = leer_excel_columna1(fp)
    comunas = [d[0] for d in datos]
    os.remove(fp)
    
    with open(cache, "w", encoding="utf-8") as f:
        json.dump(comunas, f, ensure_ascii=False)
    logging.info(f"Comunas encontradas: {len(comunas)}")
    return comunas


# ======================================================
# PASO 2: Obtener mapeo comuna → fiscalía principal
# ======================================================

def obtener_mapeo_fiscalia(base_dir):
    """
    Descarga cruce comuna-fiscalía y determina la fiscalía principal
    de cada comuna (la que tiene más casos).
    """
    cache = os.path.join(base_dir, "cache_comuna_fiscalia.json")
    if os.path.exists(cache):
        with open(cache, "r", encoding="utf-8") as f:
            mapeo = json.load(f)
        logging.info(f"Mapeo fiscalía cargado de caché: {len(mapeo)} comunas")
        return mapeo

    logging.info("Descargando cruce comuna-fiscalía...")
    params = FIXED_PARAMS.copy()
    params.update({"ANNO": "ALL", "MES_NOMBRE": "ALL",
                   "Tabular_Fila": "'GOLD DIM_COMUNA'[GLS_COMUNA]-'GOLD DIM_FISCALIA'[GLS_FISCALIA_NORM]"})
    fp = os.path.join(base_dir, "_tmp_cf.xlsx")
    if not descargar_excel(params, fp):
        return {}
    
    datos = leer_excel_columna1(fp)
    os.remove(fp)
    
    # Parsear: cada fila es "COMUNA-Fiscalía" con un conteo
    # El separador correcto es rfind('-') ya que las comunas son mayúsculas sin guiones
    # y las fiscalías empiezan con mayúscula seguida de minúsculas
    comuna_fisc_count = defaultdict(lambda: defaultdict(int))
    
    for combined, cantidad in datos:
        # Buscar el último guión seguido de una letra mayúscula y minúsculas
        # Las comunas son MAYÚSCULAS, las fiscalías tipo "Valparaiso", "Bio Bio"
        idx = combined.rfind('-')
        if idx > 0:
            comuna = combined[:idx]
            fiscalia = combined[idx+1:]
            try:
                cnt = int(cantidad)
            except (ValueError, TypeError):
                cnt = 0
            comuna_fisc_count[comuna][fiscalia] += cnt
    
    # Para cada comuna, elegir la fiscalía con más casos
    mapeo = {}
    for comuna, fiscalias in comuna_fisc_count.items():
        mejor_fisc = max(fiscalias, key=fiscalias.get)
        mapeo[comuna] = mejor_fisc
    
    with open(cache, "w", encoding="utf-8") as f:
        json.dump(mapeo, f, ensure_ascii=False, indent=2)
    
    logging.info(f"Mapeo creado: {len(mapeo)} comunas → fiscalía principal")
    # Verificación rápida
    for c in ["CONCON", "SANTIAGO", "ANTOFAGASTA", "VALPARAISO", "TALCA"]:
        if c in mapeo:
            logging.info(f"  Verificación: {c} → {mapeo[c]}")
    return mapeo


# ======================================================
# PASO 3: Parsear cruce delito-comuna
# ======================================================

def parsear_delito_comuna(filepath, comunas_set):
    """
    Parsea Excel con Tabular_Fila = delito-comuna.
    Separa usando rfind('-') — funciona porque las comunas son
    MAYÚSCULAS sin guiones y los delitos usan el último guión como separador.
    """
    datos = []
    # Crear set para validación
    comunas_upper = set(c.upper() for c in comunas_set)
    
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if row[0] is None and not header_found:
            continue
        if not header_found:
            header_found = True
            continue
        if row[0] is None:
            continue
        
        combined = str(row[0]).strip()
        try:
            cantidad = int(row[1]) if row[1] is not None else 0
        except (ValueError, TypeError):
            cantidad = 0
        
        # Separar por rfind('-')
        idx = combined.rfind('-')
        if idx > 0:
            delito = combined[:idx].strip()
            comuna = combined[idx+1:].strip()
            datos.append((delito, comuna, cantidad))
        else:
            datos.append((combined, "DESCONOCIDO", cantidad))
    
    wb.close()
    return datos


# ======================================================
# MAIN
# ======================================================

def main():
    parser = argparse.ArgumentParser(description="Descarga relacional Fiscalía de Chile")
    parser.add_argument("--test", action="store_true",
                       help="Modo test: solo 2 meses de 2025")
    parser.add_argument("--solo-csv", action="store_true",
                       help="Solo consolidar CSV sin descargar")
    args = parser.parse_args()
    
    base_dir = os.path.dirname(os.path.abspath(__file__))
    log_file = setup_logging(os.path.join(base_dir, "logs"))
    
    logging.info("=" * 60)
    logging.info("  DESCARGA RELACIONAL - FISCALÍA DE CHILE")
    logging.info("  Delitos × Comuna × Fiscalía × Año × Mes")
    logging.info("=" * 60)
    
    if args.test:
        annos = [2025]
        meses = [1, 2]
        logging.info("MODO TEST: Solo 2025, meses 1-2")
    else:
        annos = list(range(ANNO_INICIO, ANNO_FIN + 1))
        meses = MESES
        logging.info(f"MODO COMPLETO: {ANNO_INICIO}-{ANNO_FIN}, meses 1-12")
    
    total = len(annos) * len(meses)
    logging.info(f"Descargas: {total} archivos")
    logging.info(f"Tamaño estimado: ~{total * 350 / 1024:.0f} MB")
    logging.info("")
    
    # ---- PASO 1 ----
    logging.info("PASO 1: Obtener lista de comunas")
    comunas = obtener_comunas(base_dir)
    if not comunas:
        logging.error("No se pudo obtener lista de comunas. Abortando.")
        return
    
    # ---- PASO 2 ----
    logging.info("PASO 2: Obtener mapeo comuna → fiscalía principal")
    mapeo = obtener_mapeo_fiscalia(base_dir)
    
    # ---- PASO 3 ----
    if not args.solo_csv:
        logging.info("PASO 3: Descargar cruce delito × comuna")
        
        excel_dir = os.path.join(base_dir, "excel_relacional")
        os.makedirs(excel_dir, exist_ok=True)
        
        progress_file = os.path.join(base_dir, "progreso_relacional.json")
        completados = cargar_progreso(progress_file)
        
        pendientes = []
        for anno in annos:
            for mes in meses:
                tid = f"rel_{anno}_{mes:02d}"
                if tid not in completados:
                    pendientes.append((tid, anno, mes))
        
        logging.info(f"  Pendientes: {len(pendientes)} / {total}")
        
        exitosos = fallidos = 0
        inicio = time.time()
        
        for i, (tid, anno, mes) in enumerate(pendientes):
            pct = ((total - len(pendientes) + i + 1) / total) * 100
            logging.info(f"[{pct:.1f}%] {anno}-{mes:02d} ({i+1}/{len(pendientes)})")
            
            params = FIXED_PARAMS.copy()
            params["ANNO"] = str(anno)
            params["MES_NOMBRE"] = str(mes)
            params["Tabular_Fila"] = "'GOLD DIM_MATERIA'[GLS_MATERIA]-'GOLD DIM_COMUNA'[GLS_COMUNA]"
            
            fp = os.path.join(excel_dir, f"delito_comuna_{anno}_{mes:02d}.xlsx")
            
            if descargar_excel(params, fp):
                exitosos += 1
                completados.add(tid)
                guardar_progreso(progress_file, completados)
                logging.info(f"  ✓ {os.path.getsize(fp)/1024:.0f} KB")
            else:
                fallidos += 1
                logging.error(f"  ✗ FALLÓ")
            
            if i < len(pendientes) - 1:
                time.sleep(PAUSA_ENTRE_DESCARGAS)
        
        logging.info(f"\nDescargas: {exitosos} ok, {fallidos} fallidos "
                    f"en {(time.time()-inicio)/60:.1f} min")
    
    # ---- PASO 4: Consolidar CSV ----
    logging.info("\nPASO 4: Consolidar CSV relacional")
    
    excel_dir = os.path.join(base_dir, "excel_relacional")
    csv_path = os.path.join(base_dir, "datos_relacional.csv")
    
    filas = 0
    with open(csv_path, "w", newline="", encoding="utf-8") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(["anno", "mes", "delito", "comuna", 
                         "fiscalia_regional", "delitos_ingresados"])
        
        for anno in annos:
            for mes in meses:
                fp = os.path.join(excel_dir, f"delito_comuna_{anno}_{mes:02d}.xlsx")
                if not os.path.exists(fp):
                    continue
                
                datos = parsear_delito_comuna(fp, comunas)
                for delito, comuna, cantidad in datos:
                    fiscalia = mapeo.get(comuna, "")
                    writer.writerow([anno, mes, delito, comuna, fiscalia, cantidad])
                    filas += 1
                
                logging.info(f"  {anno}-{mes:02d}: {len(datos):,} registros")
    
    size_mb = os.path.getsize(csv_path) / (1024 * 1024)
    logging.info(f"\n✓ CSV: {csv_path}")
    logging.info(f"  {filas:,} filas, {size_mb:.1f} MB")
    
    # Muestra
    logging.info("\nMuestra:")
    with open(csv_path, "r", encoding="utf-8") as f:
        for i, line in enumerate(f):
            if i > 10:
                break
            logging.info(f"  {line.rstrip()}")
    
    logging.info("\n¡PROCESO COMPLETADO!")


if __name__ == "__main__":
    main()
